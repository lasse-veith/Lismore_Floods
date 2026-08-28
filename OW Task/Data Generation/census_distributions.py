"""Build per-suburb ABS 2021 Census probability distributions for Lismore.

Reads three 2021 Census GCP tables (SAL = Suburb/Locality level, LGA =
Lismore local government area, used as a fallback):
  - G34  Number of motor vehicles by dwellings
  - G36  Dwelling structure
  - G41  Dwelling structure by number of bedrooms

For each suburb currently present in properties.csv, produces normalized
probability distributions (dwelling structure, bedrooms conditional on
structure, motor vehicles per dwelling), falling back to the Lismore LGA
aggregate when a suburb's SAL-level cells are missing/suppressed or its
total dwelling count is too small to trust.

Output: lismore_census_distributions.json, keyed by suburb name (uppercase,
matching the `suburb` field already used in properties.csv/geojson).

Does not touch any other pipeline step (polygon.py, gnaf_append.py,
elevation_append.py, assembly.py) or their outputs.
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent            # Data Generation
ROOT_DIR = BASE_DIR.parent                              # OW Task
OUTPUT_DIR = ROOT_DIR / "Output"
DATA_SOURCES_DIR = ROOT_DIR / "Data Sources"

PROPERTIES_CSV = OUTPUT_DIR / "properties.csv"

GCP_ROOT = DATA_SOURCES_DIR / "2021_GCP_all_for_NSW_short-header"
TEMPLATE_XLSX = GCP_ROOT / "Metadata" / "2021_GCP_Sequential_Template_R2.xlsx"
GEOG_DESC_XLSX = GCP_ROOT / "Metadata" / "2021Census_geog_desc_1st_2nd_3rd_release.xlsx"

SAL_DIR = GCP_ROOT / "2021 Census GCP All Geographies for NSW" / "SAL" / "NSW"
LGA_DIR = GCP_ROOT / "2021 Census GCP All Geographies for NSW" / "LGA" / "NSW"

G34_SAL = SAL_DIR / "2021Census_G34_NSW_SAL.csv"
G36_SAL = SAL_DIR / "2021Census_G36_NSW_SAL.csv"
G41_SAL = SAL_DIR / "2021Census_G41_NSW_SAL.csv"
G34_LGA = LGA_DIR / "2021Census_G34_NSW_LGA.csv"
G36_LGA = LGA_DIR / "2021Census_G36_NSW_LGA.csv"
G41_LGA = LGA_DIR / "2021Census_G41_NSW_LGA.csv"

LGA_NAME = "Lismore"
LGA_CODE = "LGA14850"  # resolved from 2021Census_geog_desc: ASGS_Structure='LGA', Census_Name_2021='Lismore'

OUTPUT_JSON = OUTPUT_DIR / "lismore_census_distributions.json"

MIN_DWELLINGS_FOR_SAL = 50

# --------------------------------------------------------------------------
# Resolved CSV column mapping (verified against the exact SAL/LGA CSV headers
# and cross-checked against the G34/G36/G41 sheets of the sequential
# template — see print_table_column_meanings() below for the raw dump).
# --------------------------------------------------------------------------

# G36 — Dwelling structure (Dwellings). "Dwelling structure not stated" is
# excluded and the four categories below are renormalized over themselves.
G36_STRUCTURE_COLS = {
    "separate_house": "OPDs_Separate_house_Dwellings",
    "semi_detached": "OPDs_SD_r_t_h_th_Tot_Dwgs",
    "flat_or_apartment": "OPDs_Flt_apart_Tot_Dwgs",
    "other": "OPDs_Other_dwelling_Tot_Dwgs",
}
G36_TOTAL_COL = "OPDs_Tot_OPDs_Dwellings"  # "Total occupied private dwellings" — used only for the <50 threshold check

# G34 — Number of motor vehicles per dwelling. "Not stated" is excluded.
# "Three" and "Four or more" are combined into a single "3+" bucket.
G34_VEHICLE_COLS = {
    "0": ["Num_MVs_per_dweling_0_MVs"],
    "1": ["Num_MVs_per_dweling_1_MVs"],
    "2": ["Num_MVs_per_dweling_2_MVs"],
    "3+": ["Num_MVs_per_dweling_3_MVs", "Num_MVs_per_dweling_4mo_MVs"],
}

# G41 — Dwelling structure by number of bedrooms. "None (studio/bedsitter)"
# and "Not stated" are excluded; buckets are renormalized over 1-4 and "5+"
# (5 bedrooms merged with "6 or more", since the requested schema only goes
# to "5+"). Column-name abbreviations are NOT uniform across structure types
# (e.g. the semi-detached "6 or more" suffix is "_6_m", not "_6_or_m"), so
# every column name below was copied verbatim from the actual CSV header
# rather than generated from a prefix/suffix pattern.
G41_BEDROOM_COLS = {
    "separate_house": {
        "1": ["Separate_house_NofB_1"],
        "2": ["Separate_house_NofB_2"],
        "3": ["Separate_house_NofB_3"],
        "4": ["Separate_house_NofB_4"],
        "5+": ["Separate_house_NofB_5", "Separate_house_NofB_6_or_m"],
    },
    "semi_detached": {
        "1": ["Se_d_r_or_t_h_t_Tot_NofB_1"],
        "2": ["Se_d_r_or_t_h_t_Tot_NofB_2"],
        "3": ["Se_d_r_or_t_h_t_Tot_NofB_3"],
        "4": ["Se_d_r_or_t_h_t_Tot_NofB_4"],
        "5+": ["Se_d_r_or_t_h_t_Tot_NofB_5", "Se_d_r_or_t_h_t_Tot_NofB_6_m"],
    },
    "flat_or_apartment": {
        "1": ["Flt_apart_Tot_NofB_1"],
        "2": ["Flt_apart_Tot_NofB_2"],
        "3": ["Flt_apart_Tot_NofB_3"],
        "4": ["Flt_apart_Tot_NofB_4"],
        "5+": ["Flt_apart_Tot_NofB_5", "Flt_apart_Tot_NofB_6_or_m"],
    },
    "other": {
        "1": ["Other_dwelling_NofB_1"],
        "2": ["Other_dwelling_NofB_2"],
        "3": ["Other_dwelling_NofB_3"],
        "4": ["Other_dwelling_NofB_4"],
        "5+": ["Other_dwelling_NofB_5", "Other_dwelling_NofB_6_or_m"],
    },
}


# --------------------------------------------------------------------------
# Step 1: print the column layout straight from the xlsx template, plus the
# resolved mapping above, so it can be confirmed before anything is computed
# --------------------------------------------------------------------------

def print_table_column_meanings() -> None:
    wb = openpyxl.load_workbook(TEMPLATE_XLSX, read_only=True, data_only=True)
    for sheet_name, title in (("G34", "G34 — Number of motor vehicles by dwellings"),
                               ("G36", "G36 — Dwelling structure"),
                               ("G41", "G41 — Dwelling structure by number of bedrooms")):
        print(f"\n=== {title} (raw rows from {TEMPLATE_XLSX.name}, sheet '{sheet_name}') ===")
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                print("  " + " | ".join(str(c) for c in row if c is not None))

    print("\n=== Resolved short-header CSV column mapping used by this script ===")
    print("G36 dwelling_structure_dist categories:")
    for name, col in G36_STRUCTURE_COLS.items():
        print(f"  {name:20s} <- {col}")
    print(f"  (threshold/fallback check uses total column: {G36_TOTAL_COL})")

    print("\nG34 vehicle_count_dist buckets:")
    for bucket, cols in G34_VEHICLE_COLS.items():
        print(f"  {bucket:6s} <- {' + '.join(cols)}")

    print("\nG41 bedroom_dist_by_structure buckets:")
    for structure, buckets in G41_BEDROOM_COLS.items():
        print(f"  {structure}:")
        for bucket, cols in buckets.items():
            print(f"    {bucket:6s} <- {' + '.join(cols)}")


# --------------------------------------------------------------------------
# Suburb list (from properties.csv, so this stays correct if coords.txt and
# the resulting suburb set ever change) and SAL code resolution
# --------------------------------------------------------------------------

def load_suburb_counts() -> dict[str, int]:
    counts: dict[str, int] = {}
    with PROPERTIES_CSV.open(newline="") as f:
        for row in csv.DictReader(f):
            suburb = row["suburb"]
            counts[suburb] = counts.get(suburb, 0) + 1
    return counts


def load_valid_nsw_sal_codes() -> set[str]:
    with G36_SAL.open(newline="") as f:
        return {row["SAL_CODE_2021"] for row in csv.DictReader(f)}


def resolve_sal_codes(suburb_names: list[str]) -> dict[str, str | None]:
    """Map each suburb name to its NSW SAL code via the ABS geography
    descriptor workbook. Handles disambiguating suffixes like "Lismore
    (NSW)" vs "Lismore (Vic.)" by restricting candidates to codes that
    actually appear in the NSW SAL data files."""
    valid_codes = load_valid_nsw_sal_codes()
    wb = openpyxl.load_workbook(GEOG_DESC_XLSX, read_only=True, data_only=True)
    ws = wb["2021_ASGS_Non_ABS_Structures"]

    targets = {name.upper(): name for name in suburb_names}
    candidates: dict[str, list[str]] = {name: [] for name in suburb_names}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] != "SAL":
            continue
        code, name = row[1], row[3]
        if not code or not name:
            continue
        normalized = re.sub(r"\s*\([^)]*\)\s*$", "", str(name)).strip().upper()
        code_str = str(code)
        if normalized in targets and code_str in valid_codes:
            candidates[targets[normalized]].append(code_str)

    resolved: dict[str, str | None] = {}
    for name, codes in candidates.items():
        resolved[name] = codes[0] if len(codes) == 1 else None
    return resolved


# --------------------------------------------------------------------------
# CSV loading + numeric parsing (ABS uses ".." for confidentiality-suppressed
# cells, and blank for not-applicable/genuinely missing rows)
# --------------------------------------------------------------------------

def load_rows_by_code(path: Path, code_col: str) -> dict[str, dict[str, str]]:
    with path.open(newline="") as f:
        return {row[code_col]: row for row in csv.DictReader(f)}


def to_number(value: str | None) -> float | None:
    if value is None:
        return None
    text = value.strip()
    if text == "" or text == "..":
        return None
    try:
        return float(text)
    except ValueError:
        return None


# --------------------------------------------------------------------------
# Step 2: per-suburb fallback decision
# --------------------------------------------------------------------------

def check_missing_or_suppressed(
    g34_row: dict[str, str] | None,
    g36_row: dict[str, str] | None,
    g41_row: dict[str, str] | None,
) -> list[str]:
    reasons: list[str] = []
    if g34_row is None:
        reasons.append("G34 row missing")
    if g36_row is None:
        reasons.append("G36 row missing")
    if g41_row is None:
        reasons.append("G41 row missing")
    if g34_row is None or g36_row is None or g41_row is None:
        return reasons  # can't check individual cells without the rows

    total_dwellings = to_number(g36_row.get(G36_TOTAL_COL))
    if total_dwellings is None:
        reasons.append(f"G36 {G36_TOTAL_COL} suppressed/empty")
    elif total_dwellings < MIN_DWELLINGS_FOR_SAL:
        reasons.append(f"G36 total dwellings ({total_dwellings:.0f}) < {MIN_DWELLINGS_FOR_SAL}")

    for col in G36_STRUCTURE_COLS.values():
        if to_number(g36_row.get(col)) is None:
            reasons.append(f"G36 {col} suppressed/empty")

    for cols in G34_VEHICLE_COLS.values():
        for col in cols:
            if to_number(g34_row.get(col)) is None:
                reasons.append(f"G34 {col} suppressed/empty")

    for buckets in G41_BEDROOM_COLS.values():
        for cols in buckets.values():
            for col in cols:
                if to_number(g41_row.get(col)) is None:
                    reasons.append(f"G41 {col} suppressed/empty")

    return reasons


# --------------------------------------------------------------------------
# Step 3: normalize raw counts into probability distributions
# --------------------------------------------------------------------------

def normalize(counts: dict[str, float]) -> dict[str, float]:
    total = sum(counts.values())
    if total <= 0:
        return {k: 0.0 for k in counts}
    return {k: v / total for k, v in counts.items()}


def build_dwelling_structure_dist(g36_row: dict[str, str]) -> dict[str, float]:
    counts = {name: (to_number(g36_row.get(col)) or 0.0) for name, col in G36_STRUCTURE_COLS.items()}
    return normalize(counts)


def build_vehicle_count_dist(g34_row: dict[str, str]) -> dict[str, float]:
    counts = {
        bucket: sum(to_number(g34_row.get(col)) or 0.0 for col in cols)
        for bucket, cols in G34_VEHICLE_COLS.items()
    }
    return normalize(counts)


def build_bedroom_dist_by_structure(g41_row: dict[str, str]) -> dict[str, dict[str, float]]:
    result = {}
    for structure, buckets in G41_BEDROOM_COLS.items():
        counts = {
            bucket: sum(to_number(g41_row.get(col)) or 0.0 for col in cols)
            for bucket, cols in buckets.items()
        }
        result[structure] = normalize(counts)
    return result


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main() -> None:
    print_table_column_meanings()

    suburb_counts = load_suburb_counts()
    suburb_names = sorted(suburb_counts)
    print(f"\nSuburbs found in {PROPERTIES_CSV.name}: {len(suburb_names)}")

    sal_codes = resolve_sal_codes(suburb_names)

    g34_sal = load_rows_by_code(G34_SAL, "SAL_CODE_2021")
    g36_sal = load_rows_by_code(G36_SAL, "SAL_CODE_2021")
    g41_sal = load_rows_by_code(G41_SAL, "SAL_CODE_2021")
    g34_lga = load_rows_by_code(G34_LGA, "LGA_CODE_2021")
    g36_lga = load_rows_by_code(G36_LGA, "LGA_CODE_2021")
    g41_lga = load_rows_by_code(G41_LGA, "LGA_CODE_2021")

    lga_g34_row = g34_lga.get(LGA_CODE)
    lga_g36_row = g36_lga.get(LGA_CODE)
    lga_g41_row = g41_lga.get(LGA_CODE)
    if lga_g34_row is None or lga_g36_row is None or lga_g41_row is None:
        raise RuntimeError(f"Lismore LGA row ({LGA_CODE}) missing from LGA-level census tables")

    output: dict[str, dict] = {}
    summary_rows: list[tuple[str, str, str]] = []  # (suburb, source, detail)

    for suburb in suburb_names:
        sal_code = sal_codes.get(suburb)
        g34_row = g34_sal.get(sal_code) if sal_code else None
        g36_row = g36_sal.get(sal_code) if sal_code else None
        g41_row = g41_sal.get(sal_code) if sal_code else None

        reasons = [] if sal_code is None else check_missing_or_suppressed(g34_row, g36_row, g41_row)
        if sal_code is None:
            reasons = ["no unique NSW SAL code found for this suburb name"] + reasons

        if reasons:
            source = "LGA_fallback"
            used_g34, used_g36, used_g41 = lga_g34_row, lga_g36_row, lga_g41_row
            detail = "; ".join(reasons[:3]) + (" ..." if len(reasons) > 3 else "")
        else:
            # sal_code/g34_row/g36_row/g41_row are guaranteed non-None here:
            # reasons is only empty when sal_code resolved AND
            # check_missing_or_suppressed found all three rows present.
            assert sal_code is not None and g34_row is not None and g36_row is not None and g41_row is not None
            source = "SAL"
            used_g34, used_g36, used_g41 = g34_row, g36_row, g41_row
            detail = sal_code

        output[suburb] = {
            "source": source,
            "sal_code": sal_code,
            "fallback_reasons": reasons,
            "property_count": suburb_counts[suburb],
            "total_dwellings_used": to_number(used_g36.get(G36_TOTAL_COL)),
            "dwelling_structure_dist": build_dwelling_structure_dist(used_g36),
            "bedroom_dist_by_structure": build_bedroom_dist_by_structure(used_g41),
            "vehicle_count_dist": build_vehicle_count_dist(used_g34),
        }
        summary_rows.append((suburb, source, detail))

    OUTPUT_JSON.write_text(json.dumps(output, indent=2))

    print(f"\n=== Wrote {OUTPUT_JSON.name} ===")
    print("\n=== Summary: source used per suburb ===")
    print(f"{'Suburb':20s} {'Properties':>10s}  {'Source':13s} Detail")
    for suburb, source, detail in summary_rows:
        print(f"{suburb:20s} {suburb_counts[suburb]:>10d}  {source:13s} {detail}")

    fallback_suburbs = [s for s, source, _ in summary_rows if source == "LGA_fallback"]
    print(f"\nSuburbs using LGA fallback: {fallback_suburbs if fallback_suburbs else 'none'}")


if __name__ == "__main__":
    main()
